# orphans_screen.py
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import sqlite3
from typing import Optional, Tuple
from ttkbootstrap.widgets import DateEntry

# Ù‚Ø§Ø¦Ù…Ø© Ù…Ø­Ø§ÙØ¸Ø§Øª Ø§Ù„Ø¹Ø±Ø§Ù‚
GOVERNORATES = [
    "",
    "Ø¨ØºØ¯Ø§Ø¯",
    "Ù†ÙŠÙ†ÙˆÙ‰",
    "Ø§Ù„Ø¨ØµØ±Ø©",
    "Ø§Ù„Ù†Ø¬Ù",
    "ÙƒØ±Ø¨Ù„Ø§Ø¡",
    "Ø§Ù„Ø£Ù†Ø¨Ø§Ø±",
    "Ø¯ÙŠØ§Ù„Ù‰",
    "ØµÙ„Ø§Ø­ Ø§Ù„Ø¯ÙŠÙ†",
    "ÙˆØ§Ø³Ø·",
    "Ù…ÙŠØ³Ø§Ù†",
    "Ø°ÙŠ Ù‚Ø§Ø±",
    "Ø§Ù„Ù…Ø«Ù†Ù‰",
    "Ø¨Ø§Ø¨Ù„",
    "ÙƒØ±ÙƒÙˆÙƒ",
    "Ø¯Ù‡ÙˆÙƒ",
    "Ø£Ø±Ø¨ÙŠÙ„",
    "Ø§Ù„Ø³Ù„ÙŠÙ…Ø§Ù†ÙŠØ©",
    "Ø§Ù„Ù‚Ø§Ø¯Ø³ÙŠØ©",
]


class OrphansScreen(ttk.Frame):
    def __init__(self, parent, conn: sqlite3.Connection):
        super().__init__(parent)

        self.conn = conn

        # ØªÙ‚Ø³ÙŠÙ… Ø§Ù„ÙØ±ÙŠÙ… Ø¥Ù„Ù‰ Ø¹Ù…ÙˆØ¯ÙŠÙ†: ÙŠØ³Ø§Ø± (Ù‚Ø§Ø¦Ù…Ø© Ø§Ù„Ø£ÙŠØªØ§Ù…) + ÙŠÙ…ÙŠÙ† (Ø§Ù„ÙÙˆØ±Ù…)
        self.rowconfigure(0, weight=1)
        self.columnconfigure(0, weight=3)   # Ø§Ù„Ø¬Ø¯ÙˆÙ„
        self.columnconfigure(1, weight=2)   # Ø§Ù„ÙÙˆØ±Ù…

        # Ù…ØªØºÙŠØ±Ø§Øª Ø¨ÙŠØ§Ù†Ø§Øª Ø§Ù„ÙŠØªÙŠÙ…
        self.var_id = tk.StringVar()
        self.var_name = tk.StringVar()
        self.var_gender = tk.StringVar()
        self.var_age = tk.StringVar()
        self.var_governorate = tk.StringVar()
        self.var_address = tk.StringVar()
        self.var_guardian_name = tk.StringVar()
        self.var_guardian_phone = tk.StringVar()
        self.var_status = tk.StringVar(value="ÙØ¹Ù‘Ø§Ù„")

        # Ù…ØªØºÙŠØ±Ø§Øª Ø¨ÙŠØ§Ù†Ø§Øª Ø§Ù„ÙƒÙØ§Ù„Ø©
        self.var_monthly_amount = tk.StringVar()
        self.var_sponsorship_status = tk.StringVar(value="ÙØ¹Ù‘Ø§Ù„Ø©")
        self.var_sponsorship_start = tk.StringVar()

        # Ù…ØªØºÙŠØ±Ø§Øª Ø§Ù„ÙÙ„ØªØ±Ø© / Ø§Ù„Ø¨Ø­Ø«
        self.filter_name = tk.StringVar()
        self.filter_governorate = tk.StringVar()
        self.filter_status = tk.StringVar()

        self.create_widgets()
        self.load_orphans()

    # ==========================
    #   Ø¯ÙˆØ§Ù„ Ù…Ø³Ø§Ø¹Ø¯Ø©
    # ==========================
    def _apply_right_tag(self, text_widget: tk.Text):
        """Ù…Ø­Ø§Ø°Ø§Ø© Ø§Ù„Ù†Øµ ÙÙŠ Text Ù„Ù„ÙŠÙ…ÙŠÙ† Ø¨Ø§Ø³ØªØ®Ø¯Ø§Ù… tag (Ø¨Ø¯Ù„ justify)."""
        text_widget.tag_configure("right", justify="right")
        text_widget.tag_add("right", "1.0", "end")

    def _fmt_amount(self, value) -> str:
        """ØªÙ†Ø³ÙŠÙ‚ Ù…Ø¨Ù„Øº (Ù„Ù„Ù…Ø¨Ù„Øº Ø§Ù„Ø´Ù‡Ø±ÙŠ Ø¹Ù†Ø¯ Ø§Ù„ØªØµØ¯ÙŠØ±)."""
        if value is None:
            return "0"
        try:
            v = float(value)
        except (TypeError, ValueError):
            return str(value)
        if v.is_integer():
            return str(int(v))
        return f"{v:.2f}"
    

    def _validate_int(self, value: str) -> bool:
        """Ø§Ù„Ø³Ù…Ø§Ø­ ÙÙ‚Ø· Ø¨Ø£Ø±Ù‚Ø§Ù… ØµØ­ÙŠØ­Ø© Ø£Ùˆ ÙØ±Ø§Øº ÙÙŠ Ø­Ù‚Ù„ Ø§Ù„Ø¹Ù…Ø±."""
        return value == "" or value.isdigit()


    # ==========================
    #   Ø§Ù„ÙˆØ§Ø¬Ù‡Ø©
    # ==========================

    def create_widgets(self):
        """Ø¥Ù†Ø´Ø§Ø¡ Ø§Ù„ÙˆØ§Ø¬Ù‡Ø§Øª (Ø§Ù„Ø¬Ø¯ÙˆÙ„ + Ø§Ù„ÙÙˆØ±Ù…)"""

        # Ø£ÙˆØ§Ù…Ø± ØªØ­Ù‚Ù‚ Ù„Ù„Ø£Ø±Ù‚Ø§Ù… (Ù…Ø«Ù„Ø§Ù‹ Ù„Ù„Ø¹Ù…Ø±)
        vcmd_int = (self.register(self._validate_int), "%P")

        # Ø³ØªØ§ÙŠÙ„ Ù„Ù„Ù€ Treeview
        style = ttk.Style(self)
        style.configure("Orphans.Treeview", rowheight=28)

        # ====== ÙŠØ³Ø§Ø±: Ù‚Ø§Ø¦Ù…Ø© Ø§Ù„Ø£ÙŠØªØ§Ù… ======
        table_frame = ttk.LabelFrame(
            self, text="Ù‚Ø§Ø¦Ù…Ø© Ø§Ù„Ø£ÙŠØªØ§Ù…", padding=5, labelanchor="ne", style="Card.TLabelframe"
        )
        table_frame.grid(row=0, column=0, sticky="nsew", padx=10, pady=10)

        # ----- Ø´Ø±ÙŠØ· Ø§Ù„Ø¨Ø­Ø« ÙˆØ§Ù„ÙÙ„ØªØ±Ø© ÙÙˆÙ‚ Ø§Ù„Ø¬Ø¯ÙˆÙ„ -----
        # Ø¥Ø·Ø§Ø± Ø®Ø§Ø±Ø¬ÙŠ Ù…Ø¹ Ø¹Ù†ÙˆØ§Ù† Ø£Ù†ÙŠÙ‚
        search_frame = ttk.LabelFrame(table_frame, text="Ø£Ø¯ÙˆØ§Øª Ø§Ù„Ø¨Ø­Ø« ÙˆØ§Ù„ÙÙ„ØªØ±Ø©", padding=10)
        search_frame.pack(fill="x", padx=5, pady=(0, 10))

        # ØªÙ‚Ø³ÙŠÙ… Ø§Ù„Ø¥Ø·Ø§Ø±: Ø¹Ù…ÙˆØ¯ Ù„Ù„Ø£Ø²Ø±Ø§Ø± (ÙŠØ³Ø§Ø±) ÙˆØ¹Ù…ÙˆØ¯ Ù„Ù„Ù…Ø¯Ø®Ù„Ø§Øª (ÙŠÙ…ÙŠÙ†)
        search_frame.columnconfigure(0, weight=0)  # Ø§Ù„Ø£Ø²Ø±Ø§Ø±
        search_frame.columnconfigure(1, weight=1)  # Ù…Ø³Ø§ÙØ© ÙØ§Ø±ØºØ© Ù…Ø±Ù†Ø©
        search_frame.columnconfigure(2, weight=0)  # Ø§Ù„Ù…Ø¯Ø®Ù„Ø§Øª

        # --- Ø£ÙˆÙ„Ø§Ù‹: Ø§Ù„Ø­Ù‚ÙˆÙ„ (ØªØ¸Ù‡Ø± Ø¹Ù„Ù‰ Ø§Ù„ÙŠÙ…ÙŠÙ†) ---
        # Ù†Ø³ØªØ®Ø¯Ù… Frame Ø¯Ø§Ø®Ù„ÙŠ Ù„ØªØ±ØªÙŠØ¨ Ø§Ù„Ø­Ù‚ÙˆÙ„ Ø¨Ø¬Ø§Ù†Ø¨ Ø¨Ø¹Ø¶Ù‡Ø§
        inputs_frame = ttk.Frame(search_frame)
        inputs_frame.grid(row=0, column=2, sticky="e")

        # 1. Ø¨Ø­Ø« Ø¨Ø§Ù„Ø§Ø³Ù…
        ttk.Label(inputs_frame, text="Ø§Ù„Ø§Ø³Ù… :\u200f").pack(side="right", padx=(5, 0))
        entry_search_name = ttk.Entry(
            inputs_frame,
            textvariable=self.filter_name,
            width=20,
            justify="right"
        )
        entry_search_name.pack(side="right", padx=(10, 0))
        # Ø±Ø¨Ø· Ø²Ø± Enter Ø¨Ø§Ù„Ø¨Ø­Ø«
        entry_search_name.bind("<Return>", lambda e: self.apply_filters())

        # 2. Ø§Ù„Ù…Ø­Ø§ÙØ¸Ø©
        ttk.Label(inputs_frame, text="Ø§Ù„Ù…Ø­Ø§ÙØ¸Ø© :\u200f").pack(side="right", padx=(5, 0))
        combo_filter_gov = ttk.Combobox(
            inputs_frame,
            textvariable=self.filter_governorate,
            values=GOVERNORATES,
            width=12,
            state="readonly",
            justify="right"
        )
        combo_filter_gov.pack(side="right", padx=(10, 0))
        combo_filter_gov.current(0)

        # 3. Ø­Ø§Ù„Ø© Ø§Ù„ÙŠØªÙŠÙ…
        ttk.Label(inputs_frame, text="Ø§Ù„Ø­Ø§Ù„Ø© :\u200f").pack(side="right", padx=(5, 0))
        combo_filter_status = ttk.Combobox(
            inputs_frame,
            textvariable=self.filter_status,
            values=["", "ÙØ¹Ù‘Ø§Ù„", "Ù…ÙˆÙ‚ÙˆÙ", "Ù…Ù†Ø³Ø­Ø¨"],
            width=10,
            state="readonly",
            justify="right"
        )
        combo_filter_status.pack(side="right", padx=(0, 0))
        combo_filter_status.current(0)

        # --- Ø«Ø§Ù†ÙŠØ§Ù‹: Ø§Ù„Ø£Ø²Ø±Ø§Ø± (ØªØ¸Ù‡Ø± Ø¹Ù„Ù‰ Ø§Ù„ÙŠØ³Ø§Ø±) ---
        btns_frame = ttk.Frame(search_frame)
        btns_frame.grid(row=0, column=0, sticky="w")

        # Ø²Ø± Ù…Ø³Ø­ (Ø¨ØªØµÙ…ÙŠÙ… Outline ÙˆÙ„ÙˆÙ† Ø«Ø§Ù†ÙˆÙŠ)
        btn_clear_filter = ttk.Button(
            btns_frame,
            text="Ù…Ø³Ø­ âœ–",
            bootstyle="secondary-outline",
            width=8,
            command=self.clear_filters,
        )
        btn_clear_filter.pack(side="left", padx=(0, 5))

        # Ø²Ø± ØªØ·Ø¨ÙŠÙ‚ (Ø¨ØªØµÙ…ÙŠÙ… ØµÙ„Ø¨ ÙˆÙ„ÙˆÙ† Ø£Ø³Ø§Ø³ÙŠ)
        btn_apply_filter = ttk.Button(
            btns_frame,
            text="Ø¨Ø­Ø« ğŸ”",
            bootstyle="primary",
            width=8,
            command=self.apply_filters,
        )
        btn_apply_filter.pack(side="left")

        # Ø´Ø±ÙŠØ· ØªÙ…Ø±ÙŠØ±
        self.tree_scroll_y = ttk.Scrollbar(table_frame, orient="vertical")
        self.tree_scroll_y.pack(side="right", fill="y")

        # Ø¬Ø¯ÙˆÙ„ Ø§Ù„Ø£ÙŠØªØ§Ù… (Ù†Ø¹Ø±Ø¶Ù‡ Ù…Ù† Ø§Ù„ÙŠÙ…ÙŠÙ† Ù„Ù„ÙŠØ³Ø§Ø± Ø¨Ø§Ø³ØªØ®Ø¯Ø§Ù… displaycolumns)
        self.tree = ttk.Treeview(
            table_frame,
            columns=("id", "name", "age", "governorate", "status", "monthly_amount"),
            show="headings",
            displaycolumns=("monthly_amount", "status", "governorate", "age", "name", "id"),
            yscrollcommand=self.tree_scroll_y.set,
            style="Orphans.Treeview",
        )
        self.tree.pack(expand=True, fill="both")

        self.tree_scroll_y.config(command=self.tree.yview)

        # Ø¹Ù†Ø§ÙˆÙŠÙ† Ø§Ù„Ø£Ø¹Ù…Ø¯Ø©
        self.tree.heading("id", text="Ø±Ù‚Ù… ")
        self.tree.heading("name", text="Ø§Ù„Ø§Ø³Ù…")
        self.tree.heading("age", text="Ø§Ù„Ø¹Ù…Ø±")
        self.tree.heading("governorate", text="Ø§Ù„Ù…Ø­Ø§ÙØ¸Ø©")
        self.tree.heading("status", text="Ø­Ø§Ù„Ø© Ø§Ù„ÙŠØªÙŠÙ…")
        self.tree.heading("monthly_amount", text="Ø§Ù„Ù…Ø¨Ù„Øº Ø§Ù„Ø´Ù‡Ø±ÙŠ")

        # Ø­Ø¬Ù… Ø§Ù„Ø£Ø¹Ù…Ø¯Ø©
        self.tree.column("id", width=60, anchor="center")
        self.tree.column("name", width=180, anchor="center")
        self.tree.column("age", width=60, anchor="center")
        self.tree.column("governorate", width=100, anchor="center")
        self.tree.column("status", width=80, anchor="center")
        self.tree.column("monthly_amount", width=100, anchor="center")

        # Ø­Ø¯Ø« Ø§Ø®ØªÙŠØ§Ø± ØµÙ
        self.tree.bind("<<TreeviewSelect>>", self.on_tree_select)

        # ---- Ø£Ø²Ø±Ø§Ø± Ø§Ù„ØªØµØ¯ÙŠØ± ØªØ­Øª Ø§Ù„Ø¬Ø¯ÙˆÙ„ ----
        export_frame = ttk.Frame(table_frame)
        export_frame.pack(fill="x", padx=5, pady=(5, 0))

        btn_excel = ttk.Button(
            export_frame,
            text="Excel ØªØµØ¯ÙŠØ± Ø§Ù„Ø£ÙŠØªØ§Ù… Ø¥Ù„Ù‰",
            width=22,
            command=self.export_to_excel,
        )
        btn_excel.pack(side="right", padx=(5, 0))

        btn_pdf = ttk.Button(
            export_frame,
            text="PDF ØªØµØ¯ÙŠØ± Ø§Ù„Ø£ÙŠØªØ§Ù… Ø¥Ù„Ù‰",
            width=22,
            command=self.export_to_pdf,
        )
        btn_pdf.pack(side="right", padx=(5, 0))

        # ====== ÙŠÙ…ÙŠÙ†: Ù†Ù…ÙˆØ°Ø¬ Ø¨ÙŠØ§Ù†Ø§Øª Ø§Ù„ÙŠØªÙŠÙ… ÙˆØ§Ù„ÙƒÙØ§Ù„Ø© ======
        form_frame = ttk.LabelFrame(
            self, text="Ø¨ÙŠØ§Ù†Ø§Øª Ø§Ù„ÙŠØªÙŠÙ… ÙˆØ§Ù„ÙƒÙØ§Ù„Ø©", padding=10, labelanchor="ne", style="Card.TLabelframe"
        )

        form_frame.grid(row=0, column=1, sticky="nsew", padx=(0, 10), pady=10)

        form_frame.columnconfigure(0, weight=1)
        form_frame.columnconfigure(1, weight=0)

        row_idx = 0

        def add_row(label_text: str, widget):
            """ÙŠØ¶ÙŠÙ Ø­Ù‚Ù„ + Ø¹Ù†ÙˆØ§Ù†Ù‡ ÙÙŠ ØµÙ ÙˆØ§Ø­Ø¯ØŒ Ù…Ø¹ Ø§Ù„Ù†Ù‚Ø·ØªÙŠÙ† Ø¨Ø¹Ø¯ Ø§Ù„Ù†Øµ."""
            nonlocal row_idx
            widget.grid(row=row_idx, column=0, sticky="w", pady=4, padx=(0, 5))
            ttk.Label(form_frame, text=label_text + " :\u200f").grid(
                row=row_idx, column=1, sticky="e", pady=4, padx=(5, 0)
            )
            row_idx += 1

        # Ø±Ù‚Ù… Ø§Ù„ÙŠØªÙŠÙ… (Ù‚Ø±Ø§Ø¡Ø© ÙÙ‚Ø·)
        entry_id = ttk.Entry(
            form_frame, textvariable=self.var_id, state="readonly", width=10, justify="right"
        )
        add_row("Ø±Ù‚Ù… Ø§Ù„Ù…Ù„Ù", entry_id)

        # Ø§Ù„Ø§Ø³Ù… Ø§Ù„Ø«Ù„Ø§Ø«ÙŠ
        entry_name = tk.Entry(
            form_frame, 
            textvariable=self.var_name, 
            width=35, 
            justify="right", 
            font=("Arial", 11)
        )
        add_row("Ø§Ù„Ø§Ø³Ù… Ø§Ù„Ø«Ù„Ø§Ø«ÙŠ", entry_name)

        # Ø§Ù„Ø¬Ù†Ø³
        gender_combo = ttk.Combobox(
            form_frame,
            textvariable=self.var_gender,
            values=["", "Ø°ÙƒØ±", "Ø£Ù†Ø«Ù‰"],
            width=12,
            state="readonly",
            justify="right",
        )
        gender_combo.current(0)
        add_row("Ø§Ù„Ø¬Ù†Ø³", gender_combo)

        # Ø§Ù„Ø¹Ù…Ø±
        entry_age = ttk.Entry(
            form_frame,
            textvariable=self.var_age,
            width=10,
            justify="right",
            validate="key",
            validatecommand=vcmd_int,  # ÙŠØ³Ù…Ø­ ÙÙ‚Ø· Ø¨Ø§Ù„Ø£Ø±Ù‚Ø§Ù… Ø£Ùˆ ÙØ±Ø§Øº
        )
        add_row("Ø§Ù„Ø¹Ù…Ø± (Ø¨Ø§Ù„Ø³Ù†ÙˆØ§Øª)", entry_age)

        # Ø§Ù„Ù…Ø­Ø§ÙØ¸Ø© - Ù‚Ø§Ø¦Ù…Ø© Ù…Ù†Ø³Ø¯Ù„Ø©
        gov_combo = ttk.Combobox(
            form_frame,
            textvariable=self.var_governorate,
            values=GOVERNORATES,
            width=20,
            state="readonly",
            justify="right",
        )
        gov_combo.current(0)
        add_row("Ø§Ù„Ù…Ø­Ø§ÙØ¸Ø©", gov_combo)

        # Ø§Ù„Ø¹Ù†ÙˆØ§Ù† Ø§Ù„ÙƒØ§Ù…Ù„
        entry_address = ttk.Entry(
            form_frame, textvariable=self.var_address, width=35, justify="right"
        )
        add_row("Ø§Ù„Ø¹Ù†ÙˆØ§Ù† Ø§Ù„ÙƒØ§Ù…Ù„", entry_address)

        # Ø§Ø³Ù… ÙˆÙ„ÙŠ Ø§Ù„Ø£Ù…Ø±
        entry_guard_name = ttk.Entry(
            form_frame, textvariable=self.var_guardian_name, width=35, justify="right"
        )
        add_row("Ø§Ø³Ù… ÙˆÙ„ÙŠ Ø§Ù„Ø£Ù…Ø±", entry_guard_name)

        # Ù‡Ø§ØªÙ ÙˆÙ„ÙŠ Ø§Ù„Ø£Ù…Ø±
        entry_guard_phone = ttk.Entry(
            form_frame,
            textvariable=self.var_guardian_phone,
            width=20,
            justify="right",
            validate="key",
            validatecommand=vcmd_int,   # â¬…ï¸ Ø£Ø±Ù‚Ø§Ù… ÙÙ‚Ø·
        )
        add_row("Ù‡Ø§ØªÙ ÙˆÙ„ÙŠ Ø§Ù„Ø£Ù…Ø±", entry_guard_phone)

        # Ø­Ø§Ù„Ø© Ø§Ù„ÙŠØªÙŠÙ…
        status_combo = ttk.Combobox(
            form_frame,
            textvariable=self.var_status,
            values=["ÙØ¹Ù‘Ø§Ù„", "Ù…ÙˆÙ‚ÙˆÙ", "Ù…Ù†Ø³Ø­Ø¨"],
            width=12,
            state="readonly",
            justify="right",
        )
        status_combo.current(0)
        add_row("Ø­Ø§Ù„Ø© Ø§Ù„ÙŠØªÙŠÙ…", status_combo)

        # Ù…Ù„Ø§Ø­Ø¸Ø§Øª Ø¹Ø§Ù…Ø©
        self.notes_text = tk.Text(form_frame, width=35, height=3, wrap="word")
        self.notes_text.grid(row=row_idx, column=0, sticky="w", pady=4, padx=(0, 5))
        self._apply_right_tag(self.notes_text)
        self.notes_text.bind(
            "<KeyRelease>", lambda e: self._apply_right_tag(self.notes_text)
        )
        ttk.Label(form_frame, text="Ù…Ù„Ø§Ø­Ø¸Ø§Øª (Ø¹Ù† Ø§Ù„ÙˆØ¶Ø¹ Ø§Ù„Ø¹Ø§Ù…) :\u200f").grid(
            row=row_idx, column=1, sticky="e", pady=4, padx=(5, 0)
        )
        row_idx += 1

        # ÙØ§ØµÙ„
        ttk.Separator(form_frame, orient="horizontal").grid(
            row=row_idx, column=0, columnspan=2, sticky="ew", pady=10
        )
        row_idx += 1

        ttk.Label(form_frame, text="Ø¨ÙŠØ§Ù†Ø§Øª Ø§Ù„ÙƒÙØ§Ù„Ø©", font=("Arial", 10, "bold")).grid(
            row=row_idx, column=1, columnspan=2, sticky="e", pady=(0, 5)
        )
        row_idx += 1

        # Ø§Ù„Ù…Ø¨Ù„Øº Ø§Ù„Ø´Ù‡Ø±ÙŠ
        entry_amount = ttk.Entry(
            form_frame,
            textvariable=self.var_monthly_amount,
            width=12,
            justify="right",
            validate="key",
            validatecommand=vcmd_int,   # â¬…ï¸ Ø£Ø±Ù‚Ø§Ù… ÙÙ‚Ø·
        )
        add_row("Ø§Ù„Ù…Ø¨Ù„Øº Ø§Ù„Ø´Ù‡Ø±ÙŠ", entry_amount)

        # ØªØ§Ø±ÙŠØ® Ø¨Ø¯Ø¡ Ø§Ù„ÙƒÙØ§Ù„Ø© - Ø¨Ø§Ø³ØªØ®Ø¯Ø§Ù… ØªÙ‚ÙˆÙŠÙ… ttkbootstrap
        # Ù„Ø§Ø­Ø¸: ØºÙŠØ±Ù†Ø§ date_pattern Ø¥Ù„Ù‰ dateformat ÙˆØ§Ø³ØªØ®Ø¯Ù…Ù†Ø§ ØµÙŠØºØ© %Y-%m-%d
        entry_start = DateEntry(
            form_frame,
            width=15,
            dateformat='%Y-%m-%d',
        )
        # Ù†Ø±Ø¨Ø· Ø§Ù„Ù…ØªØºÙŠØ± ÙˆØ§Ù„Ù†Øµ Ù„Ù„ÙŠÙ…ÙŠÙ† ÙŠØ¯ÙˆÙŠØ§Ù‹ Ù„Ø£Ù† DateEntry Ù‡Ù†Ø§ Ø¹Ø¨Ø§Ø±Ø© Ø¹Ù† Ø­Ø§ÙˆÙŠØ©
        entry_start.entry.configure(textvariable=self.var_sponsorship_start, justify="right")
        
        # Ù†Ø³ØªØ®Ø¯Ù… pack Ø£Ùˆ grid Ù„Ù„Ø­Ø§ÙˆÙŠØ© Ù†ÙØ³Ù‡Ø§
        # Ù„Ø§Ø­Ø¸: Ø§Ù„Ø¯Ø§Ù„Ø© add_row ØªØªÙˆÙ‚Ø¹ widgetØŒ Ù„Ø°Ø§ Ø³Ù†Ù…Ø±Ø± Ù„Ù‡Ø§ entry_start
        add_row("ØªØ§Ø±ÙŠØ® Ø¨Ø¯Ø¡ Ø§Ù„ÙƒÙØ§Ù„Ø©", entry_start)

        # Ø­Ø§Ù„Ø© Ø§Ù„ÙƒÙØ§Ù„Ø©
        sponsorship_status_combo = ttk.Combobox(
            form_frame,
            textvariable=self.var_sponsorship_status,
            values=["ÙØ¹Ù‘Ø§Ù„Ø©", "Ù…ÙˆÙ‚ÙˆÙØ©", "Ù…Ù†ØªÙ‡ÙŠØ©"],
            width=12,
            state="readonly",
            justify="right",
        )
        sponsorship_status_combo.current(0)
        add_row("Ø­Ø§Ù„Ø© Ø§Ù„ÙƒÙØ§Ù„Ø©", sponsorship_status_combo)

        # Ù…Ù„Ø§Ø­Ø¸Ø§Øª Ø§Ù„ÙƒÙØ§Ù„Ø©
        self.sponsorship_notes_text = tk.Text(
            form_frame, width=35, height=3, wrap="word"
        )
        self.sponsorship_notes_text.grid(
            row=row_idx, column=0, sticky="w", pady=4, padx=(0, 5)
        )
        self._apply_right_tag(self.sponsorship_notes_text)
        self.sponsorship_notes_text.bind(
            "<KeyRelease>", lambda e: self._apply_right_tag(self.sponsorship_notes_text)
        )
        ttk.Label(form_frame, text="Ù…Ù„Ø§Ø­Ø¸Ø§Øª Ø§Ù„ÙƒÙØ§Ù„Ø© :\u200f").grid(
            row=row_idx, column=1, sticky="e", pady=4, padx=(5, 0)
        )
        row_idx += 1

       # Ø£Ø²Ø±Ø§Ø± Ø§Ù„ØªØ­ÙƒÙ…
        btn_frame = ttk.Frame(form_frame)
        btn_frame.grid(row=row_idx, column=0, columnspan=2, pady=15)

        # 1. Ø²Ø± "Ø¬Ø¯ÙŠØ¯" (Ø§Ù„Ø£ÙˆÙ„ Ù…Ù† Ø§Ù„ÙŠÙ…ÙŠÙ† -> Ù†Ø¹Ø·ÙŠÙ‡ Ø§Ù„Ø¹Ù…ÙˆØ¯ 2)
        ttk.Button(btn_frame, text="Ø¬Ø¯ÙŠØ¯", width=10, bootstyle="secondary", command=self.clear_form).grid(
            row=0, column=2, padx=5
        )

        # 2. Ø²Ø± "Ø­ÙØ¸" (ÙÙŠ Ø§Ù„ÙˆØ³Ø· -> Ù†Ø¹Ø·ÙŠÙ‡ Ø§Ù„Ø¹Ù…ÙˆØ¯ 1)
        ttk.Button(
            btn_frame, text="Ø­ÙØ¸", width=10, bootstyle="primary", command=self.save_orphan_and_sponsorship
        ).grid(row=0, column=1, padx=5)

        # 3. Ø²Ø± "Ø­Ø°Ù" (Ø§Ù„Ø£Ø®ÙŠØ± Ù…Ù† Ø§Ù„ÙŠØ³Ø§Ø± -> Ù†Ø¹Ø·ÙŠÙ‡ Ø§Ù„Ø¹Ù…ÙˆØ¯ 0)
        ttk.Button(btn_frame, text="Ø­Ø°Ù", width=10, bootstyle="danger", command=self.delete_orphan).grid(
            row=0, column=0, padx=5
        )
    # ==========================
    #   Ø¹Ù…Ù„ÙŠØ§Øª Ù‚Ø§Ø¹Ø¯Ø© Ø§Ù„Ø¨ÙŠØ§Ù†Ø§Øª
    # ==========================

    def load_orphans(self):
        """ØªØ­Ù…ÙŠÙ„ Ø§Ù„Ø£ÙŠØªØ§Ù… Ù…Ø¹ Ø¨ÙŠØ§Ù†Ø§Øª Ø§Ù„ÙƒÙØ§Ù„Ø© Ø¥Ù„Ù‰ Ø§Ù„Ø¬Ø¯ÙˆÙ„ (Ù…Ø¹ ØªØ·Ø¨ÙŠÙ‚ Ø§Ù„ÙÙ„ØªØ±Ø© Ø¥Ù† ÙˆØ¬Ø¯Øª)"""
        for row in self.tree.get_children():
            self.tree.delete(row)

        cursor = self.conn.cursor()

        query = """
            SELECT
                o.id,
                o.name,
                o.age,
                o.governorate,
                o.status,
                COALESCE(s.monthly_amount, 0)
            FROM orphans o
            LEFT JOIN sponsorships s
                ON o.id = s.orphan_id
                AND (s.status = 'ÙØ¹Ù‘Ø§Ù„Ø©' OR s.status IS NULL)
            WHERE 1=1
        """
        params = []

        # ØªØ·Ø¨ÙŠÙ‚ Ø§Ù„ÙÙ„Ø§ØªØ±
        name_filter = self.filter_name.get().strip()
        gov_filter = self.filter_governorate.get().strip()
        status_filter = self.filter_status.get().strip()

        if name_filter:
            query += " AND o.name LIKE ?"
            params.append(f"%{name_filter}%")

        if gov_filter:
            query += " AND o.governorate = ?"
            params.append(gov_filter)

        if status_filter:
            query += " AND o.status = ?"
            params.append(status_filter)

        query += " ORDER BY o.id ASC"

        cursor.execute(query, params)
        rows = cursor.fetchall()

        # Ø§Ù„ØªØ¹Ø¯ÙŠÙ„: Ù†Ø³ØªØ®Ø¯Ù… enumerate Ù„Ø¹Ù…Ù„ Ø¹Ø¯Ø§Ø¯ (i) ÙŠØ¨Ø¯Ø£ Ù…Ù† 1
        for i, row in enumerate(rows, start=1):
            real_id = row[0]  # Ù‡Ø°Ø§ Ù‡Ùˆ Ø§Ù„Ù€ ID Ø§Ù„Ø­Ù‚ÙŠÙ‚ÙŠ Ù…Ù† Ù‚Ø§Ø¹Ø¯Ø© Ø§Ù„Ø¨ÙŠØ§Ù†Ø§Øª
            
            # Ù†Ø¬Ù‡Ø² Ø§Ù„Ù‚ÙŠÙ… Ù„Ù„Ø¹Ø±Ø¶: Ù†Ø¶Ø¹ Ø§Ù„Ø¹Ø¯Ø§Ø¯ (i) ÙÙŠ Ø£ÙˆÙ„ Ø¹Ù…ÙˆØ¯ Ø¨Ø¯Ù„Ø§Ù‹ Ù…Ù† Ø§Ù„Ù€ ID Ø§Ù„Ø­Ù‚ÙŠÙ‚ÙŠ
            # row[1] Ù‡Ùˆ Ø§Ù„Ø§Ø³Ù…ØŒ row[2] Ø§Ù„Ø¹Ù…Ø±... Ø¥Ù„Ø®
            display_values = (i, row[1], row[2], row[3], row[4], row[5])
            
            # Ù†Ø³ØªØ®Ø¯Ù… real_id ÙƒÙ…Ø¹Ø±Ù Ù„Ù„ØµÙ (iid) Ù„Ù†Ø³ØªØ±Ø¬Ø¹Ù‡ Ø¹Ù†Ø¯ Ø§Ù„Ù†Ù‚Ø±
            self.tree.insert("", tk.END, iid=str(real_id), values=display_values)

    def get_orphan_by_id(self, orphan_id: int) -> Optional[Tuple]:
        cursor = self.conn.cursor()
        cursor.execute("SELECT * FROM orphans WHERE id = ?", (orphan_id,))
        return cursor.fetchone()

    def get_sponsorship_by_orphan(self, orphan_id: int) -> Optional[Tuple]:
        cursor = self.conn.cursor()
        cursor.execute(
            """
            SELECT id, orphan_id, monthly_amount, start_date, status, notes
            FROM sponsorships
            WHERE orphan_id = ?
            ORDER BY id DESC
            LIMIT 1
            """,
            (orphan_id,),
        )
        return cursor.fetchone()

    def insert_orphan(self, data: dict) -> int:
        cursor = self.conn.cursor()
        cursor.execute(
            """
            INSERT INTO orphans
                (name, gender, age, governorate, address,
                 guardian_name, guardian_phone, status, notes)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                data["name"],
                data["gender"],
                data["age"],
                data["governorate"],
                data["address"],
                data["guardian_name"],
                data["guardian_phone"],
                data["status"],
                data["notes"],
            ),
        )
        self.conn.commit()
        return cursor.lastrowid

    def update_orphan(self, orphan_id: int, data: dict):
        cursor = self.conn.cursor()
        cursor.execute(
            """
            UPDATE orphans
            SET name = ?, gender = ?, age = ?, governorate = ?, address = ?,
                guardian_name = ?, guardian_phone = ?, status = ?, notes = ?
            WHERE id = ?
            """,
            (
                data["name"],
                data["gender"],
                data["age"],
                data["governorate"],
                data["address"],
                data["guardian_name"],
                data["guardian_phone"],
                data["status"],
                data["notes"],
                orphan_id,
            ),
        )
        self.conn.commit()

    def insert_or_update_sponsorship(self, orphan_id: int, data: dict):
        cursor = self.conn.cursor()
        cursor.execute(
            "SELECT id FROM sponsorships WHERE orphan_id = ? ORDER BY id DESC LIMIT 1",
            (orphan_id,),
        )
        row = cursor.fetchone()
        if row:
            sponsorship_id = row[0]
            cursor.execute(
                """
                UPDATE sponsorships
                SET monthly_amount = ?, start_date = ?, status = ?, notes = ?
                WHERE id = ?
                """,
                (
                    data["monthly_amount"],
                    data["start_date"],
                    data["status"],
                    data["notes"],
                    sponsorship_id,
                ),
            )
        else:
            cursor.execute(
                """
                INSERT INTO sponsorships
                    (orphan_id, monthly_amount, start_date, status, notes)
                VALUES (?, ?, ?, ?, ?)
                """,
                (
                    orphan_id,
                    data["monthly_amount"],
                    data["start_date"],
                    data["status"],
                    data["notes"],
                ),
            )
        self.conn.commit()

    def delete_orphan_db(self, orphan_id: int):
        cursor = self.conn.cursor()
        cursor.execute("DELETE FROM payments WHERE orphan_id = ?", (orphan_id,))
        cursor.execute("DELETE FROM sponsorships WHERE orphan_id = ?", (orphan_id,))
        cursor.execute("DELETE FROM orphans WHERE id = ?", (orphan_id,))
        self.conn.commit()

    # ==========================
    #   Ù…Ù†Ø·Ù‚ Ø§Ù„ÙˆØ§Ø¬Ù‡Ø©
    # ==========================

    def clear_form(self):
        """ØªÙØ±ÙŠØº Ø­Ù‚ÙˆÙ„ Ø§Ù„ÙÙˆØ±Ù…"""
        self.var_id.set("")
        self.var_name.set("")
        self.var_gender.set("")
        self.var_age.set("")
        self.var_governorate.set("")
        self.var_address.set("")
        self.var_guardian_name.set("")
        self.var_guardian_phone.set("")
        self.var_status.set("ÙØ¹Ù‘Ø§Ù„")
        self.notes_text.delete("1.0", tk.END)

        self.var_monthly_amount.set("")
        self.var_sponsorship_status.set("ÙØ¹Ù‘Ø§Ù„Ø©")
        self.var_sponsorship_start.set("")
        self.sponsorship_notes_text.delete("1.0", tk.END)

        self.tree.selection_remove(self.tree.selection())

    def apply_filters(self):
        """ØªØ·Ø¨ÙŠÙ‚ Ù‚ÙŠÙ… Ø§Ù„Ø¨Ø­Ø«/Ø§Ù„ÙÙ„ØªØ±Ø© Ø§Ù„Ø­Ø§Ù„ÙŠØ© Ø¹Ù„Ù‰ Ø§Ù„Ø¬Ø¯ÙˆÙ„"""
        self.load_orphans()

    def clear_filters(self):
        """Ù…Ø³Ø­ Ø§Ù„ÙÙ„Ø§ØªØ± ÙˆØ¥Ø¹Ø§Ø¯Ø© ØªØ­Ù…ÙŠÙ„ ÙƒÙ„ Ø§Ù„Ø£ÙŠØªØ§Ù…"""
        self.filter_name.set("")
        self.filter_governorate.set("")
        self.filter_status.set("")
        self.load_orphans()

    def form_orphan_data(self) -> Optional[dict]:
        """Ù‚Ø±Ø§Ø¡Ø© Ø¨ÙŠØ§Ù†Ø§Øª Ø§Ù„ÙŠØªÙŠÙ… Ù…Ù† Ø§Ù„ÙÙˆØ±Ù…"""
        name = self.var_name.get().strip()
        if not name:
            messagebox.showwarning("ØªÙ†Ø¨ÙŠÙ‡", "Ø­Ù‚Ù„ Ø§Ù„Ø§Ø³Ù… Ø§Ù„Ø«Ù„Ø§Ø«ÙŠ Ù…Ø·Ù„ÙˆØ¨.")
            return None

        age_str = self.var_age.get().strip()
        age_val: Optional[int] = None
        if age_str:
            if not age_str.isdigit():
                messagebox.showwarning("ØªÙ†Ø¨ÙŠÙ‡", "Ø­Ù‚Ù„ Ø§Ù„Ø¹Ù…Ø± ÙŠØ¬Ø¨ Ø£Ù† ÙŠÙƒÙˆÙ† Ø±Ù‚Ù…Ø§Ù‹ ØµØ­ÙŠØ­Ø§Ù‹.")
                return None
            age_val = int(age_str)

        data = {
            "name": name,
            "gender": self.var_gender.get().strip(),
            "age": age_val,
            "governorate": self.var_governorate.get().strip(),
            "address": self.var_address.get().strip(),
            "guardian_name": self.var_guardian_name.get().strip(),
            "guardian_phone": self.var_guardian_phone.get().strip(),
            "status": self.var_status.get().strip() or "ÙØ¹Ù‘Ø§Ù„",
            "notes": self.notes_text.get("1.0", tk.END).strip(),
        }
        return data

    def form_sponsorship_data(self) -> Optional[dict]:
        """Ù‚Ø±Ø§Ø¡Ø© Ø¨ÙŠØ§Ù†Ø§Øª Ø§Ù„ÙƒÙØ§Ù„Ø© Ù…Ù† Ø§Ù„ÙÙˆØ±Ù… (Ø§Ø®ØªÙŠØ§Ø±ÙŠ)"""
        amount_str = self.var_monthly_amount.get().strip()
        start_date = self.var_sponsorship_start.get().strip()
        status = self.var_sponsorship_status.get().strip() or "ÙØ¹Ù‘Ø§Ù„Ø©"
        notes = self.sponsorship_notes_text.get("1.0", tk.END).strip()

        if not amount_str and not start_date and not notes:
            # Ù„Ø§ ØªÙˆØ¬Ø¯ Ø¨ÙŠØ§Ù†Ø§Øª ÙƒÙØ§Ù„Ø© Ù…Ø¯Ø®Ù„Ø©
            return None

        try:
            amount_val = float(amount_str) if amount_str else 0.0
        except ValueError:
            messagebox.showwarning("ØªÙ†Ø¨ÙŠÙ‡", "Ø§Ù„Ù…Ø¨Ù„Øº Ø§Ù„Ø´Ù‡Ø±ÙŠ ÙŠØ¬Ø¨ Ø£Ù† ÙŠÙƒÙˆÙ† Ø±Ù‚Ù…ÙŠØ§Ù‹.")
            return None

        data = {
            "monthly_amount": amount_val,
            "start_date": start_date,
            "status": status,
            "notes": notes,
        }
        return data

    def save_orphan_and_sponsorship(self):
        """Ø­ÙØ¸ Ø£Ùˆ ØªØ­Ø¯ÙŠØ« Ø§Ù„ÙŠØªÙŠÙ… Ù…Ø¹ Ø¨ÙŠØ§Ù†Ø§Øª Ø§Ù„ÙƒÙØ§Ù„Ø©"""
        orphan_data = self.form_orphan_data()
        if not orphan_data:
            return

        sponsorship_data = self.form_sponsorship_data()
        orphan_id_str = self.var_id.get().strip()

        try:
            if orphan_id_str:
                orphan_id = int(orphan_id_str)
                self.update_orphan(orphan_id, orphan_data)
            else:
                orphan_id = self.insert_orphan(orphan_data)
                self.var_id.set(str(orphan_id))

            if sponsorship_data is not None:
                self.insert_or_update_sponsorship(orphan_id, sponsorship_data)

            self.load_orphans()
            messagebox.showinfo("Ù†Ø¬Ø§Ø­", "ØªÙ… Ø­ÙØ¸ Ø¨ÙŠØ§Ù†Ø§Øª Ø§Ù„ÙŠØªÙŠÙ… ÙˆØ§Ù„ÙƒÙØ§Ù„Ø© Ø¨Ù†Ø¬Ø§Ø­.")
        except Exception as e:
            messagebox.showerror("Ø®Ø·Ø£", f"Ø­Ø¯Ø« Ø®Ø·Ø£ Ø£Ø«Ù†Ø§Ø¡ Ø§Ù„Ø­ÙØ¸:\n{e}")

    def delete_orphan(self):
        """Ø­Ø°Ù ÙŠØªÙŠÙ… Ù…Ø­Ø¯Ø¯"""
        orphan_id_str = self.var_id.get().strip()
        if not orphan_id_str:
            messagebox.showwarning("ØªÙ†Ø¨ÙŠÙ‡", "Ù„Ù… ÙŠØªÙ… Ø§Ø®ØªÙŠØ§Ø± Ø£ÙŠ ÙŠØªÙŠÙ… Ù„Ù„Ø­Ø°Ù.")
            return

        confirm = messagebox.askyesno(
            "ØªØ£ÙƒÙŠØ¯ Ø§Ù„Ø­Ø°Ù", "Ù‡Ù„ Ø£Ù†Øª Ù…ØªØ£ÙƒØ¯ Ù…Ù† Ø­Ø°Ù Ù‡Ø°Ø§ Ø§Ù„ÙŠØªÙŠÙ… ÙˆØ¬Ù…ÙŠØ¹ Ø¨ÙŠØ§Ù†Ø§Øª ÙƒÙØ§Ù„ØªÙ‡ØŸ"
        )
        if not confirm:
            return

        try:
            self.delete_orphan_db(int(orphan_id_str))
            self.load_orphans()
            self.clear_form()
            messagebox.showinfo("Ù†Ø¬Ø§Ø­", "ØªÙ… Ø­Ø°Ù Ø§Ù„ÙŠØªÙŠÙ… ÙˆØ¨ÙŠØ§Ù†Ø§Øª Ø§Ù„ÙƒÙØ§Ù„Ø© Ø§Ù„Ù…Ø±ØªØ¨Ø·Ø© Ø¨Ù‡.")
        except Exception as e:
            messagebox.showerror("Ø®Ø·Ø£", f"Ø­Ø¯Ø« Ø®Ø·Ø£ Ø£Ø«Ù†Ø§Ø¡ Ø§Ù„Ø­Ø°Ù:\n{e}")

    def on_tree_select(self, event):
        """Ø¹Ù†Ø¯ Ø§Ø®ØªÙŠØ§Ø± ÙŠØªÙŠÙ… Ù…Ù† Ø§Ù„Ø¬Ø¯ÙˆÙ„ ØªØ­Ù…ÙŠÙ„ Ø¨ÙŠØ§Ù†Ø§ØªÙ‡ ÙÙŠ Ø§Ù„ÙÙˆØ±Ù…"""
        selected = self.tree.selection()
        if not selected:
            return

        # Ø§Ù„ØªØ¹Ø¯ÙŠÙ„: Ø§Ù„Ù€ orphan_id Ø§Ù„Ø¢Ù† Ù‡Ùˆ Ù†ÙØ³Ù‡ Ø§Ù„Ù€ selected[0] Ù„Ø£Ù†Ù†Ø§ Ø®Ø²Ù†Ù†Ø§Ù‡ ÙÙŠ Ø§Ù„Ù€ iid
        orphan_id = selected[0]

        row = self.get_orphan_by_id(int(orphan_id))
        if not row:
            return

        # row = (id, name, gender, age, governorate, address,
        #        guardian_name, guardian_phone, status, notes)
        self.var_id.set(row[0])
        self.var_name.set(row[1])
        self.var_gender.set(row[2] or "")
        self.var_age.set(str(row[3]) if row[3] is not None else "")
        self.var_governorate.set(row[4] or "")
        self.var_address.set(row[5] or "")
        self.var_guardian_name.set(row[6] or "")
        self.var_guardian_phone.set(row[7] or "")
        self.var_status.set(row[8] or "ÙØ¹Ù‘Ø§Ù„")
        self.notes_text.delete("1.0", tk.END)
        self.notes_text.insert("1.0", row[9] or "")
        self._apply_right_tag(self.notes_text)

        sponsorship = self.get_sponsorship_by_orphan(int(orphan_id))
        if sponsorship:
            # sponsorship = (id, orphan_id, monthly_amount, start_date, status, notes)
            self.var_monthly_amount.set(
                str(sponsorship[2]) if sponsorship[2] is not None else ""
            )
            self.var_sponsorship_start.set(sponsorship[3] or "")
            self.var_sponsorship_status.set(sponsorship[4] or "ÙØ¹Ù‘Ø§Ù„Ø©")
            self.sponsorship_notes_text.delete("1.0", tk.END)
            self.sponsorship_notes_text.insert("1.0", sponsorship[5] or "")
            self._apply_right_tag(self.sponsorship_notes_text)
        else:
            self.var_monthly_amount.set("")
            self.var_sponsorship_start.set("")
            self.var_sponsorship_status.set("ÙØ¹Ù‘Ø§Ù„Ø©")
            self.sponsorship_notes_text.delete("1.0", tk.END)
            self._apply_right_tag(self.sponsorship_notes_text)

        # ==========================
    #   Ø§Ù„ØªØµØ¯ÙŠØ± Ø¥Ù„Ù‰ Excel / PDF
    # ==========================

    def _collect_current_rows(self):
        """Ø¬Ù…Ø¹ Ø§Ù„ØµÙÙˆÙ Ø§Ù„Ø¸Ø§Ù‡Ø±Ø© Ø­Ø§Ù„ÙŠØ§Ù‹ ÙÙŠ Ø¬Ø¯ÙˆÙ„ Ø§Ù„Ø£ÙŠØªØ§Ù…."""
        rows = []
        for iid in self.tree.get_children():
            values = self.tree.item(iid)["values"]
            if values:
                # (id, name, age, governorate, status, monthly_amount)
                rows.append(values)
        return rows

    def export_to_excel(self):
        rows = self._collect_current_rows()
        if not rows:
            messagebox.showinfo("Ù…Ø¹Ù„ÙˆÙ…Ø©", "Ù„Ø§ ØªÙˆØ¬Ø¯ Ø¨ÙŠØ§Ù†Ø§Øª ÙÙŠ Ø§Ù„Ø¬Ø¯ÙˆÙ„ Ù„Ù„ØªØµØ¯ÙŠØ±.")
            return

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font
        except ImportError:
            messagebox.showerror(
                "Ø®Ø·Ø£",
                "Ù…ÙƒØªØ¨Ø© openpyxl ØºÙŠØ± Ù…Ø«Ø¨ØªØ©.\n\nØ±Ø¬Ø§Ø¡Ù‹ Ù†ÙÙ‘Ø° Ø§Ù„Ø£Ù…Ø± Ø§Ù„ØªØ§Ù„ÙŠ Ø¯Ø§Ø®Ù„ Ø¨ÙŠØ¦Ø© Ø§Ù„Ù…Ø´Ø±ÙˆØ¹:\n\npip install openpyxl",
            )
            return

        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Ù…Ù„ÙØ§Øª Excel", "*.xlsx")],
            initialfile="ØªÙ‚Ø±ÙŠØ±_Ø§Ù„Ø£ÙŠØªØ§Ù….xlsx",
            title="Ø­ÙØ¸ ØªÙ‚Ø±ÙŠØ± Ø§Ù„Ø£ÙŠØªØ§Ù… ÙƒÙ…Ù„Ù Excel",
        )
        if not file_path:
            return

        wb = Workbook()
        ws = wb.active
        ws.title = "Ø§Ù„Ø£ÙŠØªØ§Ù…"

        headers = ["Ø±Ù‚Ù… Ø§Ù„ÙŠØªÙŠÙ…", "Ø§Ù„Ø§Ø³Ù…", "Ø§Ù„Ø¹Ù…Ø±", "Ø§Ù„Ù…Ø­Ø§ÙØ¸Ø©",
                   "Ø­Ø§Ù„Ø© Ø§Ù„ÙŠØªÙŠÙ…", "Ø§Ù„Ù…Ø¨Ù„Øº Ø§Ù„Ø´Ù‡Ø±ÙŠ"]

        ws.append(headers)

        total_monthly = 0.0

        for (oid, name, age, gov, status, monthly) in rows:
            try:
                m_val = float(monthly or 0)
            except ValueError:
                m_val = 0.0
            total_monthly += m_val

            ws.append([
                oid,
                name,
                age if age is not None else "",
                gov,
                status,
                m_val,
            ])

        # ØµÙ Ù…Ø¬Ù…ÙˆØ¹ Ø§Ù„Ù…Ø¨Ø§Ù„Øº Ø§Ù„Ø´Ù‡Ø±ÙŠØ©
        ws.append([])
        ws.append([
            "", "", "", "Ù…Ø¬Ù…ÙˆØ¹ Ø§Ù„Ù…Ø¨Ø§Ù„Øº Ø§Ù„Ø´Ù‡Ø±ÙŠØ©",
            "",
            total_monthly,
        ])

        align_right = Alignment(horizontal="right")
        bold_font = Font(bold=True)

        # ØªØ±ÙˆÙŠØ³Ø©
        for cell in ws[1]:
            cell.font = bold_font
            cell.alignment = align_right

        # Ø¨Ù‚ÙŠØ© Ø§Ù„Ø®Ù„Ø§ÙŠØ§ ÙŠÙ…ÙŠÙ†
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.alignment = align_right

        # ØµÙ Ø§Ù„Ù…Ø¬Ù…ÙˆØ¹ Ø¨ÙˆÙ„Ø¯
        for cell in ws[ws.max_row]:
            cell.font = bold_font

        try:
            wb.save(file_path)
            messagebox.showinfo("ØªÙ…", "ØªÙ… ØªØµØ¯ÙŠØ± ØªÙ‚Ø±ÙŠØ± Ø§Ù„Ø£ÙŠØªØ§Ù… Ø¥Ù„Ù‰ Ù…Ù„Ù Excel Ø¨Ù†Ø¬Ø§Ø­.")
        except Exception as e:
            messagebox.showerror("Ø®Ø·Ø£", f"Ø­Ø¯Ø« Ø®Ø·Ø£ Ø£Ø«Ù†Ø§Ø¡ Ø­ÙØ¸ Ù…Ù„Ù Excel:\n{e}")

    def export_to_pdf(self):
        """ØªØµØ¯ÙŠØ± Ù‚Ø§Ø¦Ù…Ø© Ø§Ù„Ø£ÙŠØªØ§Ù… Ø§Ù„Ø¸Ø§Ù‡Ø±Ø© ÙÙŠ Ø§Ù„Ø¬Ø¯ÙˆÙ„ Ø¥Ù„Ù‰ PDF Ù…Ø¹ Ø¯Ø¹Ù… Ø§Ù„Ø¹Ø±Ø¨ÙŠØ©"""
        # Ø¬Ù…Ø¹ Ø§Ù„ØµÙÙˆÙ Ù…Ù† Ø¬Ø¯ÙˆÙ„ Ø§Ù„Ø£ÙŠØªØ§Ù…
        rows = []
        for iid in self.tree.get_children():
            values = self.tree.item(iid)["values"]
            # values = (id, name, age, governorate, status, monthly_amount)
            if values:
                rows.append(values)

        if not rows:
            messagebox.showinfo("Ù…Ø¹Ù„ÙˆÙ…Ø©", "Ù„Ø§ ØªÙˆØ¬Ø¯ Ø¨ÙŠØ§Ù†Ø§Øª ÙÙŠ Ø§Ù„Ø¬Ø¯ÙˆÙ„ Ù„Ù„ØªØµØ¯ÙŠØ±.")
            return

        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.pdfgen import canvas
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.ttfonts import TTFont
            import arabic_reshaper
            from bidi.algorithm import get_display
        except ImportError:
            messagebox.showerror(
                "Ø®Ø·Ø£",
                "Ù…ÙƒØªØ¨Ø§Øª reportlab / arabic_reshaper / python-bidi ØºÙŠØ± Ù…Ø«Ø¨ØªØ©.\n\n"
                "Ø±Ø¬Ø§Ø¡Ù‹ Ù†ÙÙ‘Ø° Ø§Ù„Ø£Ù…Ø±:\n"
                "pip install reportlab arabic_reshaper python-bidi",
            )
            return

        # ØªØ³Ø¬ÙŠÙ„ Ø®Ø· Ø¹Ø±Ø¨ÙŠ
        def register_arabic_font():
            font_paths = [
                r"C:\Windows\Fonts\arial.ttf",
                r"C:\Windows\Fonts\trado.ttf",
                r"C:\Windows\Fonts\Tahoma.ttf",
            ]
            for p in font_paths:
                try:
                    pdfmetrics.registerFont(TTFont("ArabicMain", p))
                    return "ArabicMain"
                except Exception:
                    continue
            return "Helvetica"

        font_name = register_arabic_font()

        # ØªØ¬Ù‡ÙŠØ² Ø§Ù„Ù†Øµ Ø§Ù„Ø¹Ø±Ø¨ÙŠ
        def ar(text: str) -> str:
            if text is None:
                return ""
            s = str(text)
            try:
                reshaped = arabic_reshaper.reshape(s)
                return get_display(reshaped)
            except Exception:
                return s

        from app_settings import get_default_export_dir
        initial_name = "ØªÙ‚Ø±ÙŠØ±_Ø§Ù„Ø£ÙŠØªØ§Ù….pdf"
        initial_dir = get_default_export_dir() or ""

        file_path = filedialog.asksaveasfilename(
            defaultextension=".pdf",
            filetypes=[("Ù…Ù„ÙØ§Øª PDF", "*.pdf")],
            initialfile=initial_name,
            initialdir=initial_dir,
            title="Ø­ÙØ¸ ØªÙ‚Ø±ÙŠØ± Ø§Ù„Ø£ÙŠØªØ§Ù… ÙƒÙ…Ù„Ù PDF",
        )
        if not file_path:
            return

        page_width, page_height = A4
        c = canvas.Canvas(file_path, pagesize=A4)

        # Ø¹Ù†ÙˆØ§Ù†
        c.setFont(font_name, 14)
        c.drawRightString(page_width - 40, page_height - 40, ar("ØªÙ‚Ø±ÙŠØ± Ø§Ù„Ø£ÙŠØªØ§Ù…"))

        # Ù…ÙˆØ§Ø¶Ø¹ Ø§Ù„Ø£Ø¹Ù…Ø¯Ø© (Ù…Ù† Ø§Ù„ÙŠÙ…ÙŠÙ† Ù„Ù„ÙŠØ³Ø§Ø±)
        y = page_height - 70
        x_id = page_width - 40
        x_name = x_id - 50
        x_age = x_name - 140
        x_gov = x_age - 50
        x_status = x_gov - 80
        x_amount = x_status - 80

        def draw_header():
            nonlocal y
            c.setFont(font_name, 9)
            c.drawRightString(x_id, y, ar("Ø§Ù„Ø±Ù‚Ù…"))
            c.drawRightString(x_name, y, ar("Ø§Ø³Ù… Ø§Ù„ÙŠØªÙŠÙ…"))
            c.drawRightString(x_age, y, ar("Ø§Ù„Ø¹Ù…Ø±"))
            c.drawRightString(x_gov, y, ar("Ø§Ù„Ù…Ø­Ø§ÙØ¸Ø©"))
            c.drawRightString(x_status, y, ar("Ø­Ø§Ù„Ø© Ø§Ù„ÙŠØªÙŠÙ…"))
            c.drawRightString(x_amount, y, ar("Ø§Ù„Ù…Ø¨Ù„Øº Ø§Ù„Ø´Ù‡Ø±ÙŠ"))
            y -= 15
            c.line(40, y + 5, page_width - 40, y + 5)

        # Ø±Ø³Ù… Ø§Ù„Ù‡ÙŠØ¯Ø± Ø£ÙˆÙ„ Ù…Ø±Ø©
        draw_header()
        c.setFont(font_name, 9)
        y -= 10

        # ØªØ¹Ø¨Ø¦Ø© Ø§Ù„ØµÙÙˆÙ
        for (oid, name, age, gov, status, monthly_amount) in rows:
            if y < 60:
                # ØµÙØ­Ø© Ø¬Ø¯ÙŠØ¯Ø©
                c.showPage()
                c.setFont(font_name, 14)
                c.drawRightString(page_width - 40, page_height - 40, ar("ØªÙ‚Ø±ÙŠØ± Ø§Ù„Ø£ÙŠØªØ§Ù…"))
                y = page_height - 70
                draw_header()
                c.setFont(font_name, 9)
                y -= 10

            c.drawRightString(x_id, y, str(oid))
            c.drawRightString(x_name, y, ar(name))
            c.drawRightString(
                x_age, y, str(age) if age not in (None, "") else ""
            )
            c.drawRightString(x_gov, y, ar(gov))
            c.drawRightString(x_status, y, ar(status))
            c.drawRightString(x_amount, y, str(monthly_amount or 0))
            y -= 14

        c.showPage()
        try:
            c.save()
            messagebox.showinfo("ØªÙ…", "ØªÙ… ØªØµØ¯ÙŠØ± ØªÙ‚Ø±ÙŠØ± Ø§Ù„Ø£ÙŠØªØ§Ù… Ø¥Ù„Ù‰ Ù…Ù„Ù PDF Ø¨Ù†Ø¬Ø§Ø­.")
        except Exception as e:
            messagebox.showerror("Ø®Ø·Ø£", f"Ø­Ø¯Ø« Ø®Ø·Ø£ Ø£Ø«Ù†Ø§Ø¡ Ø­ÙØ¸ Ù…Ù„Ù PDF:\n{e}")

