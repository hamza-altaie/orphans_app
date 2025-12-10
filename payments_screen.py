# payments_screen.py
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import sqlite3
from typing import Optional, Tuple
from datetime import datetime
from ttkbootstrap.widgets import DateEntry

from app_settings import get_currency_symbol, get_default_export_dir


class PaymentsScreen(ttk.Frame):
    def __init__(self, parent, conn: sqlite3.Connection):
        super().__init__(parent)

        self.conn = conn

        # متغيرات عامة
        self.var_month = tk.StringVar(value=str(datetime.now().month))
        self.var_year = tk.StringVar(value=str(datetime.now().year))
        self.var_selected_payment_id = tk.StringVar()
        self.var_pay_paid = tk.StringVar()
        self.var_pay_status = tk.StringVar()
        self.var_pay_date = tk.StringVar()

        self.pay_tree = None
        self.pay_notes_text = None

        # متغيرات الفلترة / البحث
        self.filter_name = tk.StringVar()
        self.filter_orphan_id = tk.StringVar()
        self.filter_status = tk.StringVar()

        # متغيرات المجاميع (تُعرض كنص)
        self.total_required = tk.StringVar(value="0")
        self.total_paid = tk.StringVar(value="0")
        self.total_remaining = tk.StringVar(value="0")

        # مجاميع رقمية داخلية للاستخدام في التصدير
        self._sum_required = 0.0
        self._sum_paid = 0.0
        self._sum_remaining = 0.0

        self.create_widgets()

    # ==========================
    #   دوال تحقق للأرقام
    # ==========================
    def _validate_int(self, value: str) -> bool:
        """السماح فقط بأرقام صحيحة أو فراغ."""
        return value == "" or value.isdigit()

    def _validate_float(self, value: str) -> bool:
        """السماح فقط بأرقام عشرية أو فراغ."""
        if value == "":
            return True
        try:
            float(value)
            return True
        except ValueError:
            return False

    def _fmt_amount(self, value: float) -> str:
        """تنسيق المبلغ: إذا كان عدد صحيح يظهر بدون فاصل عشري، غير ذلك برقمين عشريين."""
        if value is None:
            return "0"
        v = float(value)
        if v.is_integer():
            return str(int(v))
        return f"{v:.2f}"

    # ==========================
    #   الواجهة
    # ==========================

    def create_widgets(self):
        # ستايل بسيط للفريمات (كروت)
        style = ttk.Style(self)
        style.configure("Card.TLabelframe", padding=8)
        style.configure("Card.TLabelframe.Label", font=("Arial", 10, "bold"))

        # أوامر التحقق
        vcmd_int = (self.register(self._validate_int), "%P")
        vcmd_float = (self.register(self._validate_float), "%P")

        # سطر أعلى لاختيار الشهر والسنة + أزرار التحكم
        top_frame = ttk.Frame(self)
        top_frame.pack(fill="x", padx=10, pady=10)

        ttk.Label(top_frame, text="السنة:\u200f").pack(side="right", padx=(0, 5))
        entry_year = ttk.Entry(
            top_frame,
            textvariable=self.var_year,
            width=8,
            validate="key",
            validatecommand=vcmd_int,   # أرقام فقط
        )
        entry_year.pack(side="right", padx=(0, 15))

        ttk.Label(top_frame, text="الشهر:\u200f").pack(side="right", padx=(0, 5))
        combo_month = ttk.Combobox(
            top_frame,
            textvariable=self.var_month,
            values=[str(i) for i in range(1, 12 + 1)],
            width=5,
            state="readonly",
        )
        combo_month.pack(side="right", padx=(0, 15))

        ttk.Button(
            top_frame,
            text="تحميل الدفعات",
            command=self.load_payments_clicked,
            width=15,
        ).pack(side="right", padx=(0, 10))

        ttk.Button(
            top_frame,
            text="إنشاء دفعات الشهر",
            command=self.create_payments_clicked,
            width=18,
        ).pack(side="right", padx=(0, 10))

        # إطار رئيسي: يسار جدول، يمين فورم تحديث (باستخدام grid)
        main_frame = ttk.Frame(self)
        main_frame.pack(expand=True, fill="both", padx=10, pady=(0, 10))

        # توزيع الأعمدة: الجدول ياخذ 3 أضعاف عرض الفورم
        main_frame.columnconfigure(0, weight=3)   # الجدول
        main_frame.columnconfigure(1, weight=1)   # الفورم
        main_frame.rowconfigure(0, weight=1)

        # ===== يسار: جدول الدفعات =====
        left_frame = ttk.LabelFrame(
            main_frame,
            text="دفعات الأيتام للشهر المحدد",
            padding=5,
            labelanchor="ne",
            style="Card.TLabelframe",
        )
        left_frame.grid(row=0, column=0, sticky="nsew", padx=(0, 10))

        # ----- شريط البحث والفلترة فوق الجدول -----
        search_frame = ttk.Frame(left_frame)
        search_frame.pack(fill="x", padx=5, pady=(0, 5))

        # بحث بالاسم
        ttk.Label(search_frame, text="بحث بالاسم (يحتوي على):\u200f").grid(
            row=0, column=5, sticky="e", padx=3, pady=2
        )
        entry_search_name = ttk.Entry(
            search_frame,
            textvariable=self.filter_name,
            width=18,
            justify="right",
        )
        entry_search_name.grid(row=0, column=4, sticky="e", padx=3, pady=2)

        # رقم الملف
        ttk.Label(search_frame, text="رقم الملف:\u200f").grid(
            row=0, column=3, sticky="e", padx=3, pady=2
        )
        entry_search_oid = ttk.Entry(
            search_frame,
            textvariable=self.filter_orphan_id,
            width=10,
            justify="right",
            validate="key",
            validatecommand=vcmd_int,   # أرقام فقط
        )
        entry_search_oid.grid(row=0, column=2, sticky="e", padx=3, pady=2)

        # حالة الدفعة
        ttk.Label(search_frame, text="حالة الدفعة:\u200f").grid(
            row=1, column=5, sticky="e", padx=3, pady=2
        )
        combo_filter_status = ttk.Combobox(
            search_frame,
            textvariable=self.filter_status,
            values=["", "غير مدفوع", "مدفوع جزئياً", "مدفوع بالكامل"],
            width=16,
            state="readonly",
            justify="right",
        )
        combo_filter_status.grid(row=1, column=4, sticky="e", padx=3, pady=2)
        combo_filter_status.current(0)

        # أزرار الفلترة
        btn_apply_filter = ttk.Button(
            search_frame,
            text="تطبيق الفلترة",
            width=14,
            command=self.apply_filters,
        )
        btn_apply_filter.grid(row=0, column=0, sticky="w", padx=3, pady=2)

        btn_clear_filter = ttk.Button(
            search_frame,
            text="مسح الفلترة",
            width=14,
            command=self.clear_filters,
        )
        btn_clear_filter.grid(row=1, column=0, sticky="w", padx=3, pady=2)

        # Enter في حقل الاسم أو رقم اليتيم = تطبيق الفلترة
        entry_search_name.bind("<Return>", lambda e: self.apply_filters())
        entry_search_oid.bind("<Return>", lambda e: self.apply_filters())

        # سكرول + جدول
        scroll_y = ttk.Scrollbar(left_frame, orient="vertical")
        scroll_y.pack(side="right", fill="y")

        self.pay_tree = ttk.Treeview(
            left_frame,
            columns=("id", "orphan_id", "name", "required", "paid", "remaining", "status"),
            show="headings",
            displaycolumns=("status", "remaining", "paid", "required", "name", "orphan_id"),
            yscrollcommand=scroll_y.set,
        )
        self.pay_tree.pack(expand=True, fill="both")
        scroll_y.config(command=self.pay_tree.yview)

        self.pay_tree.heading("id", text="ID")
        self.pay_tree.heading("orphan_id", text="رقم الملف")
        self.pay_tree.heading("name", text="اسم اليتيم")
        self.pay_tree.heading("required", text="المطلوب")
        self.pay_tree.heading("paid", text="المدفوع")
        self.pay_tree.heading("remaining", text="المتبقي")
        self.pay_tree.heading("status", text="الحالة")

        self.pay_tree.column("id", width=40, anchor="center")
        self.pay_tree.column("orphan_id", width=70, anchor="center")
        self.pay_tree.column("name", width=150, anchor="center")
        self.pay_tree.column("required", width=80, anchor="center")
        self.pay_tree.column("paid", width=80, anchor="center")
        self.pay_tree.column("remaining", width=80, anchor="center")
        self.pay_tree.column("status", width=100, anchor="center")

        # ألوان حسب الحالة
        self.pay_tree.tag_configure("unpaid", background="#ffe5e5")
        self.pay_tree.tag_configure("partial", background="#fff7cc")
        self.pay_tree.tag_configure("full", background="#e5ffe5")

        self.pay_tree.bind("<<TreeviewSelect>>", self.on_payment_select)

        # ----- مجاميع تحت الجدول -----
        totals_frame = ttk.Frame(left_frame)
        totals_frame.pack(fill="x", padx=5, pady=(5, 0))

        ttk.Label(totals_frame, text="مجموع المطلوب:\u200f").grid(
            row=0, column=5, sticky="e", padx=3, pady=2
        )
        ttk.Label(totals_frame, textvariable=self.total_required).grid(
            row=0, column=4, sticky="e", padx=3, pady=2
        )

        ttk.Label(totals_frame, text="مجموع المدفوع:\u200f").grid(
            row=0, column=3, sticky="e", padx=3, pady=2
        )
        ttk.Label(totals_frame, textvariable=self.total_paid).grid(
            row=0, column=2, sticky="e", padx=3, pady=2
        )

        ttk.Label(totals_frame, text="إجمالي المتبقي:\u200f").grid(
            row=0, column=1, sticky="e", padx=3, pady=2
        )
        ttk.Label(totals_frame, textvariable=self.total_remaining).grid(
            row=0, column=0, sticky="e", padx=3, pady=2
        )

        # ----- أزرار التصدير تحت المجاميع -----
        export_frame = ttk.Frame(left_frame)
        export_frame.pack(fill="x", padx=5, pady=(5, 5))

        btn_excel = ttk.Button(
            export_frame,
            text="Excel تصدير الدفعات إلى",
            width=22,
            command=self.export_to_excel,
        )
        btn_excel.pack(side="right", padx=(5, 0))

        btn_pdf = ttk.Button(
            export_frame,
            text="PDF تصدير الدفعات إلى",
            width=22,
            command=self.export_to_pdf,
        )
        btn_pdf.pack(side="right", padx=(5, 0))

        # ===== يمين: فورم تحديث الدفعة =====
        right_frame = ttk.LabelFrame(
            main_frame,
            text="تحديث بيانات الدفعة المحددة",
            padding=10,
            labelanchor="ne",
            style="Card.TLabelframe",
        )
        # نثبته في العمود الثاني
        right_frame.grid(row=0, column=1, sticky="n", pady=0)

        right_frame.columnconfigure(0, weight=1)
        right_frame.columnconfigure(1, weight=0)

        row_idx = 0

        def add_row(label_text, widget):
            nonlocal row_idx
            widget.grid(row=row_idx, column=0, sticky="w", pady=4, padx=(0, 5))
            ttk.Label(right_frame, text=label_text + "\u200f").grid(
                row=row_idx, column=1, sticky="e", pady=4, padx=(5, 0)
            )
            row_idx += 1

        entry_pid = ttk.Entry(
            right_frame,
            textvariable=self.var_selected_payment_id,
            state="readonly",
            width=10,
        )
        add_row("رقم الدفعة:", entry_pid)

        entry_paid = ttk.Entry(
            right_frame,
            textvariable=self.var_pay_paid,
            width=12,
            validate="key",
            validatecommand=vcmd_float,   # مبلغ = رقم عشري فقط
        )
        add_row("المبلغ المدفوع:", entry_paid)

        combo_status = ttk.Combobox(
            right_frame,
            textvariable=self.var_pay_status,
            values=["غير مدفوع", "مدفوع جزئياً", "مدفوع بالكامل"],
            width=15,
            state="readonly",
        )
        add_row("حالة الدفعة:", combo_status)

        entry_date = DateEntry(
            right_frame,
            width=15,
            dateformat='%Y-%m-%d',
        )
        # ربط المتغير وتنسيق اليمين
        entry_date.entry.configure(textvariable=self.var_pay_date, justify="right")
        
        add_row("تاريخ الدفع:", entry_date)

        self.pay_notes_text = tk.Text(right_frame, width=30, height=4)
        self.pay_notes_text.grid(row=row_idx, column=0, sticky="w", pady=4, padx=(0, 5))
        ttk.Label(right_frame, text="ملاحظات:\u200f").grid(
            row=row_idx, column=1, sticky="e", pady=4, padx=(5, 0)
        )
        row_idx += 1

        ttk.Button(
            right_frame,
            text="تحديث الدفعة",
            width=15,
            command=self.update_payment_clicked,
        ).grid(row=row_idx, column=0, columnspan=2, pady=10)

    # ==========================
    #   منطق الدفعات
    # ==========================
    def prepare_payments_for_month(self, year: int, month: int):
        """
        تجهيز سجلات الدفعات لهذا الشهر:
        لكل يتيم حالته فعّال وله كفالة فعّالة (إن وجدت) ننشئ سطر دفعة إن لم يكن موجوداً.
        """
        cursor = self.conn.cursor()
        cursor.execute(
            """
            SELECT o.id, o.name, COALESCE(s.monthly_amount, 0)
            FROM orphans o
            LEFT JOIN sponsorships s
                ON o.id = s.orphan_id
                AND (s.status = 'فعّالة' OR s.status IS NULL)
            WHERE o.status = 'فعّال'
            """
        )
        orphans = cursor.fetchall()

        for orphan_id, name, monthly_amount in orphans:
            if monthly_amount is None:
                monthly_amount = 0.0

            cursor.execute(
                """
                SELECT id FROM payments
                WHERE orphan_id = ? AND month = ? AND year = ?
                """,
                (orphan_id, month, year),
            )
            row = cursor.fetchone()
            if row:
                continue

            cursor.execute(
                """
                INSERT INTO payments
                    (orphan_id, month, year, required_amount,
                     paid_amount, status, payment_date, notes)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    orphan_id,
                    month,
                    year,
                    monthly_amount,
                    0.0,
                    "غير مدفوع",
                    "",
                    "",
                ),
            )

        self.conn.commit()

    def load_payments_for_month(self, year: int, month: int):
        """تحميل الدفعات في جدول Treeview (مع تطبيق الفلترة إن وجدت)"""
        for row in self.pay_tree.get_children():
            self.pay_tree.delete(row)

        cursor = self.conn.cursor()

        query = """
            SELECT
                p.id,
                o.id,
                o.name,
                p.required_amount,
                p.paid_amount,
                p.status
            FROM payments p
            JOIN orphans o ON o.id = p.orphan_id
            WHERE p.month = ? AND p.year = ?
        """
        params = [month, year]

        name_filter = self.filter_name.get().strip()
        oid_filter = self.filter_orphan_id.get().strip()
        status_filter = self.filter_status.get().strip()

        if name_filter:
            query += " AND o.name LIKE ?"
            params.append(f"%{name_filter}%")

        if oid_filter and oid_filter.isdigit():
            query += " AND o.id = ?"
            params.append(int(oid_filter))

        if status_filter:
            query += " AND p.status = ?"
            params.append(status_filter)

        query += " ORDER BY o.id ASC"

        cursor.execute(query, params)
        rows = cursor.fetchall()

        # مجاميع
        sum_required = 0.0
        sum_paid = 0.0
        sum_remaining = 0.0

        for pid, oid, name, req, paid, status in rows:
            req = req or 0.0
            paid = paid or 0.0
            remaining = req - paid

            sum_required += req
            sum_paid += paid
            sum_remaining += remaining

            if paid <= 0:
                tag = "unpaid"
            elif paid < req:
                tag = "partial"
            else:
                tag = "full"

            self.pay_tree.insert(
                "",
                tk.END,
                values=(pid, oid, name, req, paid, remaining, status),
                tags=(tag,),
            )

        # حفظ المجاميع الرقمية
        self._sum_required = sum_required
        self._sum_paid = sum_paid
        self._sum_remaining = sum_remaining

        # تحديث المجاميع في الواجهة مع رمز العملة
        symbol = get_currency_symbol() or ""
        suffix = f" {symbol}" if symbol else ""

        self.total_required.set(self._fmt_amount(sum_required) + suffix)
        self.total_paid.set(self._fmt_amount(sum_paid) + suffix)
        self.total_remaining.set(self._fmt_amount(sum_remaining) + suffix)

    def load_payments_clicked(self):
        month_str = self.var_month.get().strip()
        year_str = self.var_year.get().strip()

        if not month_str or not year_str:
            messagebox.showwarning("تنبيه", "يرجى اختيار الشهر والسنة.")
            return

        try:
            month = int(month_str)
            year = int(year_str)
        except ValueError:
            messagebox.showwarning("تنبيه", "الشهر والسنة يجب أن يكونا أرقاماً صحيحة.")
            return

        if not (1 <= month <= 12):
            messagebox.showwarning("تنبيه", "رقم الشهر يجب أن يكون بين 1 و 12.")
            return

        self.load_payments_for_month(year, month)

        if not self.pay_tree.get_children():
            messagebox.showinfo(
                "معلومة",
                "لا توجد دفعات لهذا الشهر.\nيمكنك الضغط على 'إنشاء دفعات الشهر' لتوليدها."
            )

    def create_payments_clicked(self):
        month_str = self.var_month.get().strip()
        year_str = self.var_year.get().strip()

        if not month_str or not year_str:
            messagebox.showwarning("تنبيه", "يرجى اختيار الشهر والسنة.")
            return

        try:
            month = int(month_str)
            year = int(year_str)
        except ValueError:
            messagebox.showwarning("تنبيه", "الشهر والسنة يجب أن يكونا أرقاماً صحيحة.")
            return

        if not (1 <= month <= 12):
            messagebox.showwarning("تنبيه", "رقم الشهر يجب أن يكون بين 1 و 12.")
            return

        try:
            self.prepare_payments_for_month(year, month)
            self.load_payments_for_month(year, month)
            messagebox.showinfo("تم", "تم إنشاء/تحديث دفعات هذا الشهر بنجاح.")
        except Exception as e:
            messagebox.showerror("خطأ", f"حدث خطأ أثناء إنشاء الدفعات:\n{e}")

    def apply_filters(self):
        """تطبيق الفلاتر الحالية على الدفعات المعروضة"""
        self.load_payments_clicked()

    def clear_filters(self):
        """مسح الفلاتر وإعادة تحميل الدفعات"""
        self.filter_name.set("")
        self.filter_orphan_id.set("")
        self.filter_status.set("")
        self.load_payments_clicked()

    def on_payment_select(self, event):
        selected = self.pay_tree.selection()
        if not selected:
            return

        item = self.pay_tree.item(selected[0])
        pid, oid, name, req, paid, remaining, status = item["values"]

        self.var_selected_payment_id.set(str(pid))
        self.var_pay_paid.set(str(paid))
        self.var_pay_status.set(status if status else "")

        cursor = self.conn.cursor()
        cursor.execute(
            """
            SELECT payment_date, notes
            FROM payments
            WHERE id = ?
            """,
            (pid,),
        )
        row = cursor.fetchone()
        if row:
            self.var_pay_date.set(row[0] or "")
            self.pay_notes_text.delete("1.0", tk.END)
            self.pay_notes_text.insert("1.0", row[1] or "")
        else:
            self.var_pay_date.set("")
            self.pay_notes_text.delete("1.0", tk.END)

    def update_payment_clicked(self):
        pid_str = self.var_selected_payment_id.get().strip()
        if not pid_str:
            messagebox.showwarning("تنبيه", "لم يتم اختيار أي دفعة لتحديثها.")
            return

        try:
            pid = int(pid_str)
        except ValueError:
            messagebox.showwarning("تنبيه", "رقم الدفعة غير صحيح.")
            return

        selected = self.pay_tree.selection()
        if not selected:
            messagebox.showwarning("تنبيه", "يرجى اختيار الدفعة من الجدول أولاً.")
            return

        item = self.pay_tree.item(selected[0])
        _, _, _, req, _, _, _ = item["values"]

        paid_str = self.var_pay_paid.get().strip()
        try:
            paid = float(paid_str) if paid_str else 0.0
        except ValueError:
            messagebox.showwarning("تنبيه", "المبلغ المدفوع يجب أن يكون عدداً.")
            return

        status = self.var_pay_status.get().strip()
        if not status:
            if paid <= 0:
                status = "غير مدفوع"
            elif paid < req:
                status = "مدفوع جزئياً"
            else:
                status = "مدفوع بالكامل"

        date_str = self.var_pay_date.get().strip()
        notes = self.pay_notes_text.get("1.0", tk.END).strip()

        cursor = self.conn.cursor()
        try:
            cursor.execute(
                """
                UPDATE payments
                SET paid_amount = ?, status = ?, payment_date = ?, notes = ?
                WHERE id = ?
                """,
                (paid, status, date_str, notes, pid),
            )
            self.conn.commit()

            month = int(self.var_month.get())
            year = int(self.var_year.get())
            self.load_payments_for_month(year, month)

            messagebox.showinfo("تم", "تم تحديث بيانات الدفعة بنجاح.")
        except Exception as e:
            messagebox.showerror("خطأ", f"حدث خطأ أثناء تحديث الدفعة:\n{e}")

    # ==========================
    #   التصدير إلى Excel / PDF
    # ==========================

    def _collect_current_rows(self):
        """جمع الصفوف الظاهرة حالياً في الجدول."""
        rows = []
        for iid in self.pay_tree.get_children():
            values = self.pay_tree.item(iid)["values"]
            if values:
                # (pid, oid, name, req, paid, remaining, status)
                rows.append(values)
        return rows

    def export_to_excel(self):
        rows = self._collect_current_rows()
        if not rows:
            messagebox.showinfo("معلومة", "لا توجد بيانات في الجدول للتصدير.")
            return

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font
        except ImportError:
            messagebox.showerror(
                "خطأ",
                "مكتبة openpyxl غير مثبتة.\n\nرجاءً نفّذ الأمر التالي داخل بيئة المشروع:\n\npip install openpyxl",
            )
            return

        # اختيار مكان الحفظ
        month_str = self.var_month.get().strip()
        year_str = self.var_year.get().strip()
        initial_name = f"دفعات_الشهر_{year_str}_{month_str}.xlsx"

        from app_settings import get_default_export_dir
        initial_dir = get_default_export_dir() or ""

        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("ملفات Excel", "*.xlsx")],
            initialfile=initial_name,
            initialdir=initial_dir,
            title="حفظ تقرير الدفعات كملف Excel",
        )
        if not file_path:
            return

        wb = Workbook()
        ws = wb.active
        ws.title = "دفعات الأيتام"

        headers = ["رقم الدفعة", "رقم اليتيم", "اسم اليتيم",
                   "المطلوب", "المدفوع", "المتبقي", "حالة الدفعة"]

        ws.append(headers)

        # تعبئة البيانات
        for (pid, oid, name, req, paid, remaining, status) in rows:
            ws.append([
                pid,
                oid,
                name,
                req,
                paid,
                remaining,
                status,
            ])

        # صف المجاميع
        total_req = getattr(self, "_sum_required", 0.0)
        total_paid = getattr(self, "_sum_paid", 0.0)
        total_rem = getattr(self, "_sum_remaining", 0.0)

        ws.append([])
        ws.append([
            "", "", "المجاميع",
            total_req,
            total_paid,
            total_rem,
            "",
        ])

        align_right = Alignment(horizontal="right")
        bold_font = Font(bold=True)

        # ترويسة
        for cell in ws[1]:
            cell.font = bold_font
            cell.alignment = align_right

        # محاذاة بقية الصفوف
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.alignment = align_right

        # صف المجاميع بولد
        for cell in ws[ws.max_row]:
            cell.font = bold_font

        try:
            wb.save(file_path)
            messagebox.showinfo("تم", "تم تصدير تقرير الدفعات إلى ملف Excel بنجاح.")
        except Exception as e:
            messagebox.showerror("خطأ", f"حدث خطأ أثناء حفظ ملف Excel:\n{e}")

    def export_to_pdf(self):
        rows = self._collect_current_rows()
        if not rows:
            messagebox.showinfo("معلومة", "لا توجد بيانات في الجدول للتصدير.")
            return

        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.pdfgen import canvas
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.ttfonts import TTFont
            import arabic_reshaper
            from bidi.algorithm import get_display
        except ImportError:
            messagebox.showerror(
                "خطأ",
                "مكتبات reportlab / arabic_reshaper / python-bidi غير مثبتة.\n\n"
                "رجاءً نفّذ الأمر:\n"
                "pip install reportlab arabic_reshaper python-bidi",
            )
            return

        # تسجيل خط عربي
        def register_arabic_font():
            font_paths = [
                r"C:\Windows\Fonts\arial.ttf",
                r"C:\Windows\Fonts\trado.ttf",
                r"C:\Windows\Fonts\Tahoma.ttf",
            ]
            for p in font_paths:
                try:
                    pdfmetrics.registerFont(TTFont("ArabicMain", p))
                    return "ArabicMain"
                except Exception:
                    continue
            return "Helvetica"

        font_name = register_arabic_font()

        def ar(text: str) -> str:
            if text is None:
                return ""
            s = str(text)
            try:
                reshaped = arabic_reshaper.reshape(s)
                return get_display(reshaped)
            except Exception:
                return s

        month_str = self.var_month.get().strip()
        year_str = self.var_year.get().strip()
        initial_name = f"تقرير_دفعات_{year_str}_{month_str}.pdf"

        from app_settings import get_default_export_dir
        initial_dir = get_default_export_dir() or ""

        file_path = filedialog.asksaveasfilename(
            defaultextension=".pdf",
            filetypes=[("ملفات PDF", "*.pdf")],
            initialfile=initial_name,
            initialdir=initial_dir,
            title="حفظ تقرير الدفعات كملف PDF",
        )
        if not file_path:
            return

        page_width, page_height = A4
        c = canvas.Canvas(file_path, pagesize=A4)

        # عنوان
        c.setFont(font_name, 14)
        c.drawRightString(page_width - 40, page_height - 40,
                          ar(f"تقرير دفعات الأيتام لشهر {month_str}/{year_str}"))

        # مواضع الأعمدة (من اليمين لليسار)
        y = page_height - 70
        x_pid = page_width - 40
        x_oid = x_pid - 40
        x_name = x_oid - 140
        x_req = x_name - 70
        x_paid = x_req - 70
        x_rem = x_paid - 70
        x_status = x_rem - 70

        def draw_header():
            nonlocal y
            c.setFont(font_name, 9)
            c.drawRightString(x_pid, y, ar("رقم الدفعة"))
            c.drawRightString(x_oid, y, ar("رقم الملف"))
            c.drawRightString(x_name, y, ar("اسم اليتيم"))
            c.drawRightString(x_req, y, ar("المطلوب"))
            c.drawRightString(x_paid, y, ar("المدفوع"))
            c.drawRightString(x_rem, y, ar("المتبقي"))
            c.drawRightString(x_status, y, ar("الحالة"))
            y -= 15
            c.line(40, y + 5, page_width - 40, y + 5)

        draw_header()
        c.setFont(font_name, 9)
        y -= 10

        for (pid, oid, name, req, paid, remaining, status) in rows:
            if y < 60:
                c.showPage()
                c.setFont(font_name, 14)
                c.drawRightString(page_width - 40, page_height - 40,
                                  ar(f"تقرير دفعات الأيتام لشهر {month_str}/{year_str}"))
                y = page_height - 70
                draw_header()
                c.setFont(font_name, 9)
                y -= 10

            c.drawRightString(x_pid, y, str(pid))
            c.drawRightString(x_oid, y, str(oid))
            c.drawRightString(x_name, y, ar(name))
            c.drawRightString(x_req, y, self._fmt_amount(req))
            c.drawRightString(x_paid, y, self._fmt_amount(paid))
            c.drawRightString(x_rem, y, self._fmt_amount(remaining))
            c.drawRightString(x_status, y, ar(status))
            y -= 14

        # مجاميع في النهاية
        y -= 10
        c.line(40, y + 5, page_width - 40, y + 5)
        y -= 15

        total_req = getattr(self, "_sum_required", 0.0)
        total_paid = getattr(self, "_sum_paid", 0.0)
        total_rem = getattr(self, "_sum_remaining", 0.0)

        c.setFont(font_name, 10)
        c.drawRightString(x_name, y, ar("المجاميع"))
        c.drawRightString(x_req, y, self._fmt_amount(total_req))
        c.drawRightString(x_paid, y, self._fmt_amount(total_paid))
        c.drawRightString(x_rem, y, self._fmt_amount(total_rem))

        c.showPage()
        try:
            c.save()
            messagebox.showinfo("تم", "تم تصدير تقرير الدفعات إلى ملف PDF بنجاح.")
        except Exception as e:
            messagebox.showerror("خطأ", f"حدث خطأ أثناء حفظ ملف PDF:\n{e}")

